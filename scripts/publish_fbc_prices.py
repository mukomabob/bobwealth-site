#!/usr/bin/env python3
"""
Automates the daily FBC Securities price-sheet import for bobwealth.org.

Replicates, field-for-field, the parseSheet() / handleCurrentSheet() /
publishToGitHub() logic that already lives in markets.html's admin panel,
so market-data.json comes out identical in shape to a manual PIN-panel
upload. This script only WRITES market-data.json to the working tree —
the GitHub Actions workflow that calls it is responsible for git add/
commit/push (so the repo's own GITHUB_TOKEN handles auth, no PAT needed).

Env vars required:
  GMAIL_ADDRESS       - the Gmail account the FBC sheet arrives at
  GMAIL_APP_PASSWORD  - a Gmail "app password" for IMAP access

Exit codes:
  0  - success (including the no-op "nothing to do yet" case)
  1  - a real failure (bad credentials, unparseable sheet, etc.)
"""
import email
import imaplib
import json
import math
import os
import re
import sys
from datetime import datetime, timezone
from email.header import decode_header
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MARKET_DATA_PATH = os.path.join(REPO_ROOT, "market-data.json")
BASELINE_PATH = os.path.join(REPO_ROOT, "Price Sheet 06.01.26.xlsx")
BASELINE_FNAME = "Price_Sheet_06_01_26.xlsx"  # underscored on purpose — see sheet_date()
SHEETS_DIR = os.path.join(REPO_ROOT, "data", "fbc-sheets")  # archive for the predictions pipeline

FBC_SENDER = "FBCSECURITIESRESEARCH@fbc.co.zw"
HARARE = ZoneInfo("Africa/Harare")

EXCLUDED_SECTORS = {"Fixed Term  Bond", "Riet", "Reit", "Derivative"}
EXCLUDED_ROW_RE = re.compile(
    r"VFEX BONDS|VFEX REITS|VFEX ETF|ZSE ETF|ZSE REIT|ZSE TOP|THIS PRICE SHEET",
    re.IGNORECASE,
)


# ─── numeric helpers (mirror JS parseFloat(...)||0 / ||null quirks) ──────────
def parse_float(v):
    if v is None:
        return float("nan")
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    m = re.match(r"^[+-]?(\d+\.?\d*|\.\d+)([eE][+-]?\d+)?", s)
    if not m or m.group(0) == "":
        return float("nan")
    try:
        return float(m.group(0))
    except ValueError:
        return float("nan")


def pf_or_zero(v):
    f = parse_float(v)
    return 0 if math.isnan(f) else f


def pf_or_null(v):
    f = parse_float(v)
    if math.isnan(f) or f == 0:
        return None
    return f


# ─── date extraction (mirrors the JS regex + fallback exactly) ──────────────
def sheet_date(fname):
    m = re.search(r"(\d{2})[_-](\d{2})[_-](\d{2,4})", fname)
    if m:
        dd, mm, yy = m.group(1), m.group(2), m.group(3)
        yy2 = yy[-2:]
        return f"20{yy2}-{mm}-{dd}"
    return re.sub(r"\.(xlsx?|XLSX?)$", "", fname)


# ─── the actual sheet parser — a line-for-line port of parseSheet() ────────
def parse_sheet(raw, fname):
    hdr = -1
    for i, row in enumerate(raw):
        if row and any(str(c or "").strip().upper() == "COUNTER" for c in row):
            hdr = i
            break
    if hdr < 0:
        raise ValueError("No COUNTER header row found.")

    date = sheet_date(fname)
    mkt = "ZSE"
    rows = []

    for i in range(hdr + 1, len(raw)):
        r = raw[i]
        if not r or r[0] is None or r[0] == "":
            continue
        first = str(r[0]).strip()
        if not first:
            continue
        if "VFEX PRICE SHEET" in first.upper():
            mkt = "VFEX"
            continue
        if EXCLUDED_ROW_RE.search(first):
            continue

        def cell(idx):
            return r[idx] if idx < len(r) else None

        c1 = cell(1)
        if math.isnan(parse_float(c1)) and not isinstance(c1, str):
            continue

        sect = str(cell(5) or "").strip()
        if sect in EXCLUDED_SECTORS:
            continue

        close_px = parse_float(cell(8))
        raw13 = cell(13)
        is_susp = "SUSPEND" in str(raw13 or "").upper() or "NOT FOUND" in str(raw13 or "").upper()
        usd_price = (
            (None if math.isnan(close_px) else close_px)
            if mkt == "VFEX"
            else pf_or_null(cell(9))
        )

        rows.append({
            "counter": first,
            "isin": str(c1 or "").strip(),
            "shares": pf_or_zero(cell(2)),
            "mktCapIBR": pf_or_zero(cell(3)),
            "sector": sect or "Other",
            "open": pf_or_null(cell(7)),
            "close": None if math.isnan(close_px) else close_px,
            "usdIBR": pf_or_null(cell(9)),
            "chg": pf_or_zero(cell(11)),
            "chgPct": pf_or_zero(cell(12)),
            "vol": 0 if is_susp else pf_or_zero(cell(13)),
            "val": 0 if is_susp else pf_or_zero(cell(14)),
            "divYield": pf_or_zero(cell(15)),
            "ytd": pf_or_zero(cell(16)),
            "market": mkt,
            "suspended": is_susp,
            "usdPrice": usd_price,
        })

    return rows, date


def sheet_to_raw(xlsx_bytes):
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.worksheets[0]
    return [list(row) for row in ws.iter_rows(values_only=True)]


# ─── baseline (the fixed Jan-6-2026 sheet already committed to the repo) ────
def load_baseline():
    with open(BASELINE_PATH, "rb") as f:
        raw = sheet_to_raw(f.read())
    rows, date = parse_sheet(raw, BASELINE_FNAME)
    base = {}
    for r in rows:
        if r["usdPrice"] and r["usdPrice"] > 0:
            base[r["counter"].strip().upper()] = {
                "price": r["usdPrice"], "market": r["market"], "sector": r["sector"],
            }
    return base, date


# ─── Gmail (IMAP) ────────────────────────────────────────────────────────────
def decode_str(s):
    parts = decode_header(s or "")
    out = []
    for text, enc in parts:
        out.append(text.decode(enc or "utf-8", errors="replace") if isinstance(text, bytes) else text)
    return "".join(out)


def fetch_todays_attachment():
    """Returns (filename, bytes) for today's FBC price-sheet attachment, or (None, None)."""
    today_harare = datetime.now(HARARE).date()
    imap_date = today_harare.strftime("%d-%b-%Y")  # IMAP SINCE wants e.g. 27-Aug-2026

    addr = os.environ["GMAIL_ADDRESS"]
    app_pw = os.environ["GMAIL_APP_PASSWORD"]

    conn = imaplib.IMAP4_SSL("imap.gmail.com")
    try:
        conn.login(addr, app_pw)
        conn.select("INBOX", readonly=True)
        typ, data = conn.search(None, f'(FROM "{FBC_SENDER}" SINCE {imap_date})')
        if typ != "OK":
            raise RuntimeError(f"IMAP search failed: {typ}")
        ids = data[0].split()
        if not ids:
            return None, None

        # walk newest-first, return the first message that actually has an xlsx attachment
        for msg_id in reversed(ids):
            typ, msg_data = conn.fetch(msg_id, "(RFC822)")
            if typ != "OK":
                continue
            msg = email.message_from_bytes(msg_data[0][1])
            for part in msg.walk():
                fname = part.get_filename()
                if not fname:
                    continue
                fname = decode_str(fname)
                if fname.lower().endswith((".xlsx", ".xls")):
                    return fname, part.get_payload(decode=True)
        return None, None
    finally:
        conn.logout()


# ─── main ────────────────────────────────────────────────────────────────────
def main():
    fname, blob = fetch_todays_attachment()
    if not blob:
        print("No FBC price-sheet email found yet today — nothing to do.")
        return 0

    date_match = re.search(r"\d{2}\.\d{2}\.\d{2}", fname)
    parse_fname = f"{date_match.group(0)}.xlsx" if date_match else fname

    # Archive the raw sheet for the predictions pipeline (generate_predictions.py
    # reads every file here to rebuild the full training history). Written
    # unconditionally so a re-poll of the same day's email just overwrites it
    # with identical bytes — no-op for git.
    os.makedirs(SHEETS_DIR, exist_ok=True)
    with open(os.path.join(SHEETS_DIR, parse_fname), "wb") as f:
        f.write(blob)

    # idempotency: skip if we've already published today's date
    if os.path.exists(MARKET_DATA_PATH):
        with open(MARKET_DATA_PATH) as f:
            existing = json.load(f)
        expected_date = sheet_date(parse_fname)
        if existing.get("priceDate") == expected_date:
            print(f"market-data.json already has priceDate {expected_date} — nothing to do.")
            return 0

    raw = sheet_to_raw(blob)
    rows, price_date = parse_sheet(raw, parse_fname)
    if not rows:
        print("Parsed the sheet but found zero counters — refusing to publish an empty snapshot.", file=sys.stderr)
        return 1

    base, base_date = load_baseline()

    def inv100(r):
        b = base.get(r["counter"].strip().upper())
        if b and b["price"] > 0 and r["usdPrice"] and r["usdPrice"] > 0:
            return (r["usdPrice"] / b["price"]) * 100
        return None

    now = datetime.now(timezone.utc)
    generated = now.strftime("%Y-%m-%dT%H:%M:%S.") + f"{now.microsecond // 1000:03d}Z"

    snapshot = {
        "generated": generated,
        "priceDate": price_date,
        "baseDate": base_date,
        "counters": [
            {
                "counter": r["counter"].strip(),
                "isin": r["isin"],
                "market": r["market"],
                "sector": r["sector"],
                "shares": r["shares"],
                "mktCapIBR": r["mktCapIBR"],
                "open": r["open"],
                "close": r["close"],
                "usdIBR": r["usdIBR"],
                "chg": r["chg"],
                "chgPct": r["chgPct"],
                "vol": r["vol"],
                "val": r["val"],
                "divYield": r["divYield"],
                "ytd": r["ytd"],
                "suspended": r["suspended"],
                "usdPrice": r["usdPrice"],
                "inv100": inv100(r),
            }
            for r in rows
        ],
    }

    with open(MARKET_DATA_PATH, "w") as f:
        json.dump(snapshot, f, indent=2)
        f.write("\n")

    print(f"Wrote market-data.json — priceDate={price_date}, {len(rows)} counters, source file '{fname}'.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
